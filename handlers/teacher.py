import io
from datetime import date, datetime, timedelta
from itertools import groupby

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from aiogram import Router, F
from aiogram.filters import CommandStart
from aiogram.types import Message, BufferedInputFile, CallbackQuery
from aiogram.fsm.context import FSMContext

import database as db
from config import ADMIN_IDS, SUPERUSER_IDS
from keyboards import (
    teacher_menu_keyboard, admin_menu_keyboard,
    admin_book_manager_keyboard, book_selection_keyboard, book_delete_keyboard,
    book_edit_keyboard, report_period_keyboard
)
from states import TeacherReport, AddBook, EditBook

router = Router()


# ── Ruxsat tekshiruvi ──────────────────────────────────────────────────────────

async def check_teacher(message: Message) -> bool:
    user = await db.get_user(message.from_user.id)
    return message.from_user.id in ADMIN_IDS or message.from_user.id in SUPERUSER_IDS or (
        user is not None and user["role"] in ("teacher", "admin")
    )


async def get_menu(message: Message):
    from keyboards import superuser_menu_keyboard
    user = await db.get_user(message.from_user.id)
    if message.from_user.id in SUPERUSER_IDS:
        return superuser_menu_keyboard()
    if message.from_user.id in ADMIN_IDS or (user and user["role"] == "admin"):
        return admin_menu_keyboard()
    return teacher_menu_keyboard()


def parse_date(date_str: str) -> date | None:
    try:
        return datetime.strptime(date_str.strip(), "%d.%m.%Y").date()
    except ValueError:
        try:
            return datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None


def parse_date_range(range_str: str) -> tuple[date, date] | None:
    parts = range_str.split("-")
    if len(parts) != 2:
        return None
    start_date = parse_date(parts[0])
    end_date = parse_date(parts[1])
    if start_date and end_date:
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date
    return None


# ── Excel yaratish ─────────────────────────────────────────────────────────────

def build_excel(start_date: date, end_date: date, students: list[dict], report_type: str, class_name: str | None = None) -> io.BytesIO:
    """
    report_type: 'exercise' or 'reading'
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    type_label = "Vazifalar" if report_type == "exercise" else "Kitob o'qish"
    title_suffix = f" — {class_name}" if class_name else ""
    ws.title = f"{type_label}{title_suffix}"[:31]

    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    ALT_FILL  = PatternFill(start_color="EAF3FB", end_color="EAF3FB", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=13, color="1A3C5E")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    YES_FONT = Font(color="27AE60", bold=True)
    NO_FONT  = Font(color="E74C3C")

    is_range = start_date != end_date

    # Filter students who actually have data for this type
    if report_type == "exercise":
        filtered = [s for s in students if s["exercises"] and "Bajarmadi" not in str(s["exercises"])]
        if is_range:
             headers = ["#", "Ism", "Sinf", "Bajarilgan vazifalar", "Video yuklangan kunlar"]
        else:
             headers = ["#", "Ism", "Sinf", "Bajarilgan vazifalar", "Video"]
        col_count = 5
    else:
        filtered = [s for s in students if s["reading"] and "Bajarmadi" not in str(s["reading"]["book_name"])]
        if is_range:
             headers = ["#", "Ism", "Sinf", "Kitob", "O'qilgan betlar summasi", "Rasm yuklangan kunlar"]
        else:
             headers = ["#", "Ism", "Sinf", "Kitob", "Betlar", "Rasm"]
        col_count = 6

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    tc = ws["A1"]
    date_str = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}" if is_range else start_date.strftime('%d.%m.%Y')
    tc.value = f"{type_label} hisoboti — {date_str}{title_suffix}"
    tc.font = TITLE_FONT
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[2].height = 20

    for idx, s in enumerate(filtered, 1):
        row  = idx + 2
        fill = ALT_FILL if idx % 2 == 0 else None
        
        if report_type == "exercise":
            if is_range:
                video_val = s["exercise_video_count"]
            else:
                video_val = "✅ Ha" if s.get("exercise_video") else "❌ Yo'q"
            ex_list = ", ".join(s["exercises"])
            values = [idx, s["name"], s["class_name"], ex_list, video_val]
        else:
            if is_range:
                 photo_val = s["reading"].get("photo_count", 0)
            else:
                 photo_val = "✅ Ha" if s["reading"].get("photo_file_id") else "❌ Yo'q"
            values = [
                idx, s["name"], s["class_name"],
                s["reading"]["book_name"], s["reading"]["pages_read"], photo_val
            ]
            
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN
            cell.alignment = CENTER if col not in (4,) else LEFT
            if fill:
                cell.fill = fill
                
            # Formatting for specific columns (only for single day)
            if not is_range:
                if report_type == "exercise" and col == 5:
                    cell.font = YES_FONT if "Ha" in str(val) else NO_FONT
                elif report_type == "reading" and col == 6:
                    cell.font = YES_FONT if "Ha" in str(val) else NO_FONT
                
        ws.row_dimensions[row].height = 18

    # Set column widths
    if report_type == "exercise":
        widths = [4, 25, 14, 45, 18 if is_range else 10]
    else:
        widths = [4, 25, 14, 30, 20 if is_range else 8, 18 if is_range else 10]
        
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


async def send_report(message: Message, start_date: date, end_date: date, report_type: str):
    """
    report_type: 'exercise' or 'reading'
    """
    report_data = await db.get_report_data(start_date, end_date)
    if not report_data:
        return await message.bot.send_message(message.chat.id, "📭 Hali hech qanday o'quvchi ro'yxatdan o'tmagan.")

    label = "💪 Vazifalar" if report_type == "exercise" else "📚 Kitobxonlik"
    is_range = start_date != end_date
    date_str = f"{start_date.strftime('%d.%m.%Y')} dan {end_date.strftime('%d.%m.%Y')} gacha" if is_range else start_date.strftime('%d.%m.%Y')
    
    # 1. Generate text summary (Forward-friendly)
    text = f"{label} hisoboti — {date_str}\n\n"
    count = 0
    
    if report_type == "exercise":
        for s in report_data:
            if s["exercises"] and "Bajarmadi" not in str(s["exercises"]):
                count += 1
                ex_str = ", ".join(s["exercises"])
                if is_range:
                    video_mark = f" ({s['exercise_video_count']}ta video)" if s["exercise_video_count"] else ""
                else:
                    video_mark = " (Video ✅)" if s.get("exercise_video") else ""
                text += f"{count}. {s['name']} ({s['class_name']}): {ex_str}{video_mark}\n"
    else:
        for s in report_data:
            if s["reading"] and "Bajarmadi" not in str(s["reading"]["book_name"]):
                count += 1
                r = s["reading"]
                if is_range:
                     photo_mark = f" ({r.get('photo_count', 0)}ta rasm)" if r.get('photo_count') else ""
                else:
                     photo_mark = " (Rasm ✅)" if r.get("photo_file_id") else ""
                text += f"{count}. {s['name']} ({s['class_name']}): {r['book_name']}, {r['pages_read']} bet{photo_mark}\n"
    
    if count == 0:
        return await message.bot.send_message(message.chat.id, f"📭 {date_str} kuni uchun {label.lower()} hisoboti bo'sh.")

    text += f"\nJami: {count} nafar o'quvchi."

    # 2. Generate Excel
    excel = build_excel(start_date, end_date, report_data, report_type)
    filename = f"{report_type}_{start_date.strftime('%d%m%Y')}"
    if is_range:
         filename += f"_{end_date.strftime('%d%m%Y')}"
    filename += ".xlsx"
    
    # Send to requester
    await message.bot.send_document(
        chat_id=message.chat.id,
        document=BufferedInputFile(excel.read(), filename=filename),
        caption=text[:1024] # Telegram caption limit
    )
    if len(text) > 1024:
        await message.bot.send_message(message.chat.id, text)

    # 3. Send to class groups (only if it's a single day report, or do we want to send weekly reports to class groups too?
    # Usually, sending weekly reports to groups is also welcome.)
    class_groups = await db.get_all_class_groups()
    if class_groups:
        sorted_data = sorted(report_data, key=lambda x: x["class_name"])
        for class_name, group_iter in groupby(sorted_data, key=lambda x: x["class_name"]):
            chat_id = class_groups.get(class_name)
            if not chat_id:
                continue
            
            class_students = list(group_iter)
            # Filter class students for this report type
            if report_type == "exercise":
                filtered = [s for s in class_students if s["exercises"] and "Bajarmadi" not in str(s["exercises"])]
            else:
                filtered = [s for s in class_students if s["reading"] and "Bajarmadi" not in str(s["reading"]["book_name"])]
                
            if not filtered:
                continue
                
            class_excel = build_excel(start_date, end_date, class_students, report_type, class_name=class_name)
            
            # Detailed class caption
            c_text = f"📊 **{class_name} — {label} hisoboti**\n({date_str})\n\n"
            for i, s in enumerate(filtered, 1):
               if report_type == "exercise":
                   c_text += f"{i}. {s['name']}: {', '.join(s['exercises'])}\n"
               else:
                   c_text += f"{i}. {s['name']}: {s['reading']['book_name']}, {s['reading']['pages_read']} b.\n"
            
            c_text += f"\nJami: {len(filtered)} nafar"
            
            try:
                await message.bot.send_document(
                    chat_id=chat_id,
                    document=BufferedInputFile(class_excel.read(), filename=f"hisobot_{class_name}_{filename}"),
                    caption=c_text[:1024]
                )
                if len(c_text) > 1024:
                    await message.bot.send_message(chat_id=chat_id, text=c_text)
            except Exception as e:
                await message.bot.send_message(message.chat.id, f"⚠️ **{class_name}** guruhiga yuborishda xatolik: {e}")


# ── Report Handlers ────────────────────────────────────────────────────────────

@router.message(F.text == "💪 Vazifalar hisoboti")
async def btn_report_exercises_menu(message: Message):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")
    await message.answer("📊 Qaysi muddat uchun vazifalar hisobotini olmoqchisiz?", reply_markup=report_period_keyboard("exercise"))

@router.message(F.text == "📚 Kitoblar hisoboti")
async def btn_report_reading_menu(message: Message):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")
    await message.answer("📊 Qaysi muddat uchun kitoblar hisobotini olmoqchisiz?", reply_markup=report_period_keyboard("reading"))

@router.message(F.text == "📅 Sana bo'yicha (Vazifa)")
async def btn_report_exercises_by_date(message: Message, state: FSMContext):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")
    await btn_report_exercises_menu(message)

@router.message(F.text == "📅 Sana bo'yicha (Kitob)")
async def btn_report_reading_by_date(message: Message, state: FSMContext):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")
    await btn_report_reading_menu(message)

@router.callback_query(F.data.startswith("report_period:"))
async def cb_report_period(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    period = parts[1]
    report_type = parts[2]
    
    today = date.today()
    if period == "today":
        start_date = today
        end_date = today
        await call.message.edit_text("⏳ Hisobot tayyorlanmoqda...")
        await send_report(call.message, start_date, end_date, report_type)
    elif period == "week":
        start_date = today - timedelta(days=today.weekday()) # Monday
        end_date = today
        await call.message.edit_text("⏳ Hafta hisoboti tayyorlanmoqda...")
        await send_report(call.message, start_date, end_date, report_type)
    elif period == "month":
        start_date = today.replace(day=1) # 1st of month
        end_date = today
        await call.message.edit_text("⏳ Oy hisoboti tayyorlanmoqda...")
        await send_report(call.message, start_date, end_date, report_type)
    elif period == "custom":
        await state.set_state(TeacherReport.waiting_for_date_range)
        await state.update_data(report_type=report_type)
        await call.message.edit_text(
            "📅 Iltimos, sana yoki oraliqni yuboring.\n"
            "Masalan bitta kun uchun: **28.03.2026**\n"
            "Yoki muddat uchun: **20.03.2026-28.03.2026**"
        )
    await call.answer()


@router.message(TeacherReport.waiting_for_date_range, F.text)
async def fsm_report_date_range(message: Message, state: FSMContext):
    text = message.text.replace(" ", "")
    if "-" in text:
        dt_range = parse_date_range(text)
        if not dt_range:
            return await message.answer("❌ Noto'g'ri oraliq formati. Masalan: **20.03.2026-28.03.2026**")
        start_date, end_date = dt_range
    else:
        dt = parse_date(text)
        if not dt:
             return await message.answer("❌ Noto'g'ri sana formati. Masalan: **28.03.2026**")
        start_date = dt
        end_date = dt
    
    data = await state.get_data()
    report_type = data.get("report_type", "exercise")
    await state.clear()
    
    label = "vazifalar" if report_type == "exercise" else "kitobxonlik"
    date_str = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}" if start_date != end_date else start_date.strftime('%d.%m.%Y')
    await message.answer(f"⏳ {date_str} uchun {label} hisoboti tayyorlanmoqda...")
    await send_report(message, start_date, end_date, report_type)


# ── ⚠️ Belgilamaganlar ─────────────────────────────────────────────────────────

@router.message(F.text == "⚠️ Belgilamaganlar")
async def btn_missing(message: Message):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")

    today = date.today()
    missing = await db.get_missing_students(today)
    if not missing:
        return await message.answer(
            f"🎉 Barcha o'quvchilar bugun ({today.strftime('%d.%m.%Y')}) belgilagan!"
        )
    text = f"⚠️ **Belgilamaganlar — {today.strftime('%d.%m.%Y')}:**\n"
    current_class = None
    for s in missing:
        if s["class_name"] != current_class:
            current_class = s["class_name"]
            text += f"\n📌 **{current_class}**\n"
        text += f"  • {s['name']}\n"
    await message.answer(text)


# ── 📷 Bugungi media ─────────────────────────────────────────────────────────

@router.message(F.text == "📷 Bugungi media")
async def btn_media_today(message: Message):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")
    media = await db.get_today_media_list()
    videos = media["videos"]
    photos = media["photos"]
    if not videos and not photos:
        return await message.answer("📭 Bugun hali hech qanday video yoki rasm yuklanmagan.")
    if videos:
        await message.answer(f"🎥 **Bugungi videolar ({len(videos)} ta):**")
        for v in videos:
            caption = f"👤 {v['name']} | 🏫 {v['class_name']}"
            await message.answer_video(v["video_file_id"], caption=caption)
    if photos:
        await message.answer(f"📷 **Bugungi kitob rasmlari ({len(photos)} ta):**")
        for p in photos:
            caption = (
                f"👤 {p['name']} | 🏫 {p['class_name']}\n"
                f"📖 {p['book_name']} — {p['pages_read']} bet"
            )
            await message.answer_photo(p["photo_file_id"], caption=caption)


# ── 📚 Kitoblarni boshqarish ─────────────────────────────────────────────────

@router.message(F.text == "📚 Kitoblarni boshqarish")
async def btn_manage_books(message: Message):
    if not await check_teacher(message):
        return await message.answer("❌ Bu tugma faqat o'qituvchi/admin uchun.")
    await message.answer("📚 **Kitoblarni boshqarish menyusi**", reply_markup=admin_book_manager_keyboard())


@router.callback_query(F.data == "add_book")
async def cb_add_book_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AddBook.waiting_for_name)
    await call.message.edit_text("✏️ Yangi kitob nomini yuboring:")
    await call.answer()


@router.message(AddBook.waiting_for_name, F.text)
async def fsm_add_book_name(message: Message, state: FSMContext):
    name = message.text.strip()
    success = await db.add_book(name)
    await state.clear()
    if success:
        menu = await get_menu(message)
        await message.answer(f"✅ **{name}** kitobi qo'shildi!", reply_markup=menu)
    else:
        menu = await get_menu(message)
        await message.answer(f"⚠️ **{name}** kitobi allaqachon mavjud.", reply_markup=menu)


@router.callback_query(F.data == "list_books")
async def cb_list_books(call: CallbackQuery):
    books = await db.get_all_books()
    if not books:
        return await call.message.edit_text("📭 Kitoblar ro'yxati bo'sh.", reply_markup=admin_book_manager_keyboard())
    text = "📋 **Mavjud kitoblar:**\n\n" + "\n".join(f"• {b['name']}" for b in books)
    await call.message.edit_text(text, reply_markup=admin_book_manager_keyboard())
    await call.answer()


@router.callback_query(F.data == "list_delete_book")
async def cb_list_delete_book(call: CallbackQuery):
    books = await db.get_all_books()
    if not books:
        return await call.message.edit_text("📭 O'chirish uchun kitoblar yo'q.", reply_markup=admin_book_manager_keyboard())
    await call.message.edit_text("O'chirish uchun kitobni tanlang:", reply_markup=book_delete_keyboard(books))
    await call.answer()


@router.callback_query(F.data.startswith("delete_book:"))
async def cb_delete_book(call: CallbackQuery):
    parts = call.data.split(":")
    book_id = int(parts[1])
    book_name = parts[2]
    await db.delete_book(book_id)
    await call.message.edit_text(f"🗑 **{book_name}** kitobi o'chirildi.")
    await call.answer("O'chirildi!")


@router.callback_query(F.data == "list_edit_book")
async def cb_list_edit_book(call: CallbackQuery):
    books = await db.get_all_books()
    if not books:
        return await call.message.edit_text("📭 Tahrirlash uchun kitoblar yo'q.", reply_markup=admin_book_manager_keyboard())
    await call.message.edit_text("Tahrirlash uchun kitobni tanlang:", reply_markup=book_edit_keyboard(books))
    await call.answer()


@router.callback_query(F.data.startswith("edit_book:"))
async def cb_edit_book_select(call: CallbackQuery, state: FSMContext):
    book_id = int(call.data.split(":")[1])
    await state.update_data(edit_book_id=book_id)
    await state.set_state(EditBook.waiting_for_new_name)
    await call.message.edit_text("✏️ Kitobning yangi nomini yuboring:")
    await call.answer()


@router.message(EditBook.waiting_for_new_name, F.text)
async def fsm_edit_book_name(message: Message, state: FSMContext):
    data = await state.get_data()
    book_id = data["edit_book_id"]
    new_name = message.text.strip()
    success = await db.update_book(book_id, new_name)
    await state.clear()
    if success:
        menu = await get_menu(message)
        await message.answer(f"✅ Kitob nomi **{new_name}** ga o'zgartirildi.", reply_markup=menu)
    else:
        menu = await get_menu(message)
        await message.answer(f"⚠️ **{new_name}** nomi allaqachon mavjud yoki xatolik yuz berdi.", reply_markup=menu)
